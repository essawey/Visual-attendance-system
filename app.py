from flask import Flask, render_template, request
from groupPath import groupPath
from getRoomNum import getRoomNum
from otp import generate_otp
from start_session import start_session
from datetime import datetime
from sendEmail import checkEmail
import threading
from parallelism import parallelism_endEmail
import os
from IoT import board, lcd_command, printSYSEMSTART, lcd_init, printLCD, endLCD
import time 
import openpyxl
import re
from start_up import install_dependencies
app = Flask(__name__)
DR_EMAIL = None
COURSE_CODE = None
START_TIME = None
MAJOR = None
YEAR = None

unknow_OTP = None
REAL_OTP = generate_otp()

majors = ["Artificial Intelligence", "Computer Science", "Business", "Biotechnology", "Engineering"]
years = ["Freshman", "Sophomores", "Junior", "Senior 1", "Senior 2"]

isSessionActive = threading.Event()
isSessionActive.set()





@app.route('/', methods=['GET', 'POST'])
def index():
    printLCD("Hello !")
    if request.method == 'POST':
        global DR_EMAIL
        global COURSE_CODE
        global YEAR
        global MAJOR
        DR_EMAIL = request.form.get('dr_email')
        request_ROOM_NUM = request.form.get('room_number')
        COURSE_CODE = request.form.get('COURSE_CODE')
        YEAR = request.form.get('year')
        MAJOR = request.form.get('major')

        if not checkEmail(DR_EMAIL):
            return novaildEmail()

        if int(request_ROOM_NUM) == int(getRoomNum()):
            global START_TIME
            global isSessionActive
            START_TIME = datetime.now()
            START_TIME = str(START_TIME.strftime("%Y-%m-%d-%I-%M_%p"))

            wb = openpyxl.Workbook()
            wb.save(os.path.join(os.getcwd(),"attendence_excel.xlsx"))
            # start_session(DR_EMAIL, COURSE_CODE, REAL_OTP,
            #                     groupPath( MAJOR, YEAR, majors, years),
            #                     isSessionActive)
            sessionThread = threading.Thread(target=start_session, args=(DR_EMAIL, COURSE_CODE, REAL_OTP,
                                groupPath( MAJOR, YEAR, majors, years),
                                isSessionActive))
            sessionThread.start()
            sessionThread.join()
            return thanks()
        else:
            return noRoom()

    return render_template('index.html', majors=majors, years=years)


@app.route('/endSession', methods=['GET', 'POST'])
def endSession():
    if request.method == 'POST':
        global unknow_OTP
        global START_TIME
        global COURSE_CODE
        unknow_OTP = request.form.get('OTP')

        if unknow_OTP.strip() != REAL_OTP: # FIXME use hash and verify passwords
            return notvaild()
        
        workbook = openpyxl.load_workbook("attendence_excel.xlsx")
        worksheet = workbook['Sheet']

        worksheet.cell(worksheet.max_row+1, 1,
                        str("Attendance rate")
                    )
        
        worksheet.cell(
            worksheet.max_row, 2,
            (worksheet.max_row-2)/len(
                os.listdir(
                    os.path.join(
                        os.getcwd(),
                        groupPath(
                            MAJOR,
                            YEAR,
                            majors,
                            years
                        )
                    )
                )
            )
        )
        workbook.save('attendence_excel.xlsx')
        endLCD()
        
        START_TIME = re.sub(r'[^a-zA-Z0-9_-]', '_', str(START_TIME))
        COURSE_CODE = re.sub(r'[^a-zA-Z0-9_-]', '_', str(COURSE_CODE))

        os.rename("attendence_excel.xlsx", f"attendence_{COURSE_CODE}_{START_TIME}.xlsx")

        parallelism_endEmail(DR_EMAIL, COURSE_CODE, FILE_PATH = f"attendence_{COURSE_CODE}_{START_TIME}.xlsx")

        os.remove(f"attendence_{COURSE_CODE}_{START_TIME}.xlsx") # So user can not edit it

        # Stop
        return thanks()
    return render_template('endSession.html')


@app.route('/notvaild')
def notvaild():
    return render_template('notvaild.html')

@app.route('/noroom')
def noRoom():
    return render_template('noRoom.html')

@app.route('/thanks')
def thanks():
    isSessionActive.clear()  # This will set the isSessionActive event to False
    # Exit the IoT
    lcd_command(0x01)
    time.sleep(0.5)
    board.exit()
    return render_template('thanks.html')

@app.route('/novaildEmail')
def novaildEmail():
    return render_template('novaildEmail.html')


if __name__ == '__main__':
    # install_dependencies()
    lcd_init() # Initialization of the LCD
    printSYSEMSTART()
    app.run()