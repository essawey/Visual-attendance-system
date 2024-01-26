def start_session(DR_EMAIL,COURSE_CODE, REAL_OTP, groupPath, isSessionActive):
    import face_recognition
    import numpy as np
    import os
    import app
    import openpyxl
    from parallelism import parallelism_startEmail
    import cv2
    from datetime import datetime
    from sendEmail import is_internet_connected
    from IoT import printLCD, green_led, image_error, camera_error,image_not_found ,no_internet
    printLCD("please wait")

    # from picamera2 import Picamera2
    # picam2 = Picamera2()

    workbook = openpyxl.load_workbook("attendence_excel.xlsx")

    # Clear all cells in the current worksheet
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                cell.value = None

    if 'Sheet' not in workbook.sheetnames:
        worksheet = workbook.create_sheet('Sheet')
    else:
        worksheet = workbook['Sheet']

    worksheet['A1'] = "student ID"
    worksheet['B1'] = "Date"

    row=2
    col=1

    ###############################################################################################


    projectPath = os.getcwd()
    runPath = os.path.join(projectPath, "run")
    logPath = os.path.join(projectPath, "log")
    
    os.makedirs(logPath, exist_ok = True)
    os.makedirs(runPath, exist_ok = True)

    dataPath = os.path.join(projectPath, groupPath)

    if len(os.listdir(dataPath)) == 0:
        image_not_found()

    import glob
    image_files = glob.glob(os.path.join(dataPath, '**/*.jpg'), recursive=True)
    image_files.extend(glob.glob(os.path.join(dataPath, '**/*.png'), recursive=True))


    if len(image_files) == 0:
        image_error()

    ###############################################################################################

    # Initialize some variables
    knownEncodings = []
    knownIDs = []

    face_locations = []
    face_encodings = []
    already_attendence_taken = []

    ###############################################################################################

    for image_file in image_files:

        # load and get features
        studentImage = face_recognition.load_image_file(image_file)
        student_Face_Encoding = face_recognition.face_encodings(studentImage)[0] # only one student per pic

        # Append data
        knownEncodings.append(student_Face_Encoding)
        knownIDs.append(os.path.basename(image_file))

    ###############################################################################################

    try:
        # picam2.start_and_capture_file(os.path.join(logPath, "log.jpg"))
        pass
    except:
        pass

    if not os.path.exists(os.path.join(logPath, "log.jpg")):
        camera_error()


    if not is_internet_connected():
        no_internet()
    cap = cv2.VideoCapture(0) # port of the Built-in Cam
    parallelism_startEmail(DR_EMAIL, COURSE_CODE, REAL_OTP)
    
    printLCD("Scanning ...")

    while isSessionActive.is_set():  
        # print("while not app.END_SECTION",app.END_SECTION)

        # Grab a single frame of video
        _, frame = cap.read()
        # picam2.start_and_capture_file(os.path.join(logPath, "log.jpg"))
        cv2.imwrite(os.path.join(logPath, "log.jpg"), frame)

        frame = cv2.imread(os.path.join(logPath, "log.jpg"), cv2.COLOR_BGR2RGB)
        
        ###############################################################################################

        # Resize frame of video to 1/4 size for faster face recognition processing
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

        # Find all the faces and face encodings in the current frame of video
        face_locations = face_recognition.face_locations(small_frame, number_of_times_to_upsample = 0) # don't use photo of people in smart phones
        face_encodings = face_recognition.face_encodings(small_frame, face_locations) # N num of stdents per image


        for face_encoding in face_encodings:
            id = 'unkown'
            
            matches = face_recognition.compare_faces(knownEncodings, face_encoding, tolerance = 0.6)
            
            face_distances = face_recognition.face_distance(knownEncodings, face_encoding)

            best_match_index = np.argmin(face_distances)
            
            # See if the face is a match for the known face(s)
            if matches[best_match_index]:
                id = knownIDs[best_match_index]

            if(id[:-4] not in already_attendence_taken and id != 'unkown'):
                
                worksheet.cell(row, col, int(id[:-4]))
                col = col+1
                current_datetime = datetime.now()
                worksheet.cell(row, col, str(current_datetime.strftime("%A, %B %d, %Y %I:%M %p")))
                row = row+1
                col = 1
                workbook.save('attendence_excel.xlsx')
                printLCD(id[:-4], col = 0, row = 0)
                green_led()
                already_attendence_taken.append(id[:-4])
                
                # Display the results
                for (top, right, bottom, left) in face_locations:
                    # Scale back up face locations since the frame we detected in was scaled to 1/4 size
                    top *= 4
                    right *= 4
                    bottom *= 4
                    left *= 4

                    cv2.rectangle(frame, (left, top), (right, bottom), (36, 255, 200), 2)

                    cv2.rectangle(frame, (left, bottom - 40), (right, bottom), (36, 255, 200), cv2.FILLED)
                    font = cv2.FONT_HERSHEY_DUPLEX
                    frame = cv2.putText(frame, id[:-4], (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)
                    cv2.imwrite(os.path.join(runPath, f"{id[:-4]}.jpg"), frame)
                    